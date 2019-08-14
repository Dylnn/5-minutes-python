from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Alignment, Border, Side, Font

def excel单元格填入(sheet, value, color="FF8C69"):
    sheet.value = value
    sheet.fill = PatternFill("solid", fgColor=color) # 颜色代码：http://www.114la.com/other/rgb.htm
    # 设置数据垂直居中和水平居中
    sheet.alignment = Alignment(horizontal='center', vertical='center')#, wrap_text=True)
    border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                    top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'),
                    diagonal=Side(style='thin', color='FF000000'), diagonal_direction=0,
                    outline=Side(style='thin', color='FF000000'), vertical=Side(style='thin', color='FF000000'),
                    horizontal=Side(style='thin', color='FF000000'))
    sheet.border = border


wb = Workbook()
ws = wb.active
ws.title = '机器学习sheet'

ws = wb.create_sheet('新建了一个新的sheet')

excel单元格填入(ws['A1'], '机器学习测试结论')
ws.merge_cells('A1:B1')

data = {
    '样本数':207684,
    '分类正确':207386,
    '准确率':99.8,
    '错误率':0.14,
    '正例':1300,
    '反例':200000,
    '精确率':99.123,
    '召回率':90.456
}

i = 2
for k, v in data.items():
    excel单元格填入(ws['A' + str(i)], k, 'FFFFFF')
    excel单元格填入(ws['B' + str(i)], v, 'FFFFFF')
    i += 1

wb.save('机器学习测试.xlsx')